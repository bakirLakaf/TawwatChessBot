# -*- coding: utf-8 -*-
"""
قراءة جدول نتائج بطولة من ملف إكسل (.xlsx) — مثل الملفات المُصدَّرة من
موقع chess-results.com أو أي جدول مشابه.

يتعرّف تلقائيًا على صف الترويسة وأعمدة: الترتيب / الاسم / النقاط / التصنيف
مهما كانت تسمية الأعمدة (عربي أو إنجليزي)، ويتجاهل الصفوف الفارغة.
لا يحتاج ضبطًا يدويًا لكل بطولة — فقط أرسل الملف كما هو.
"""
import logging
import openpyxl

log = logging.getLogger(__name__)

# كلمات مفتاحية لكل عمود (تُقارَن بعد التحويل لحروف صغيرة + إزالة الفراغات)
_COL_KEYS = {
    "rank":   ["rk", "rk.", "rank", "ترتيب", "المركز", "الترتيب", "#", "no", "no."],
    "name":   ["name", "اسم", "الاسم", "player", "اللاعب", "fullname"],
    "points": ["pts", "pts.", "points", "نقاط", "النقاط", "score", "tb1"],
    "rating": ["rtg", "rtg.", "rating", "تصنيف", "التصنيف", "elo", "elo rtg"],
    "fed":    ["fed", "fed.", "federation", "دولة", "الدولة", "club", "النادي"],
}


def _norm(v):
    return str(v).strip().lower() if v is not None else ""


def _match_col(cell_value):
    """يحدّد نوع العمود من نص الترويسة، أو None إن لم يُعرف."""
    v = _norm(cell_value)
    if not v:
        return None
    for key, words in _COL_KEYS.items():
        if v in words:
            return key
    # مطابقة جزئية احتياطية (مثلاً "Pts" داخل "Pts.")
    for key, words in _COL_KEYS.items():
        for w in words:
            if w and (v.startswith(w) or w.startswith(v)):
                return key
    return None


def _find_header_row(ws, max_scan=12):
    """يبحث عن أول صف فيه عمود "الاسم" + عمود آخر على الأقل (صف الترويسة)."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            key = _match_col(ws.cell(r, c).value)
            if key and key not in mapping.values():
                mapping[c] = key
        if len(mapping) >= 2 and "name" in mapping.values():
            return r, mapping
    return None, None


def parse_excel(path):
    """يعيد قائمة dict: [{rank, name, points, rating, fed}, ...]
    من أول ورقة عمل (sheet) يجد فيها جدول نتائج صالح. يعيد [] إن لم يجد شيئًا."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        header_row, mapping = _find_header_row(ws)
        if not mapping:
            continue
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row = {}
            has_value = False
            for c, key in mapping.items():
                val = ws.cell(r, c).value
                if val not in (None, ""):
                    has_value = True
                if key not in row:  # أوّل عمود مطابق فقط لكل نوع
                    row[key] = val
            if not has_value or not row.get("name"):
                continue
            rows.append({
                "rank": str(row.get("rank") or len(rows) + 1).strip(),
                "name": str(row.get("name")).strip(),
                "points": _fmt_num(row.get("points")),
                "rating": _fmt_num(row.get("rating")),
                "fed": str(row.get("fed") or "").strip(),
            })
        if rows:
            log.info("تمّت قراءة %d لاعبًا من ملف النتائج (%s).", len(rows), ws.title)
            return rows
    return []


def _fmt_num(v):
    """ينسّق رقمًا (نقاط/تصنيف) بدون أصفار عشرية زائدة، أو '' إن كان فارغًا."""
    if v in (None, ""):
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(round(v, 1))
    return str(v).strip()
